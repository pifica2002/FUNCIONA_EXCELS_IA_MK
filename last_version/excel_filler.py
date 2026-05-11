import os
from openpyxl import load_workbook
import torch

# ============================================================
# QWEN GENERATION
# ============================================================

@torch.inference_mode()
def qwen_generate(model, processor, system_prompt, user_prompt, gen_kwargs):
    messages = [
        {"role": "system", "content": [{"type": "text", "text": system_prompt}]},
        {"role": "user", "content": [{"type": "text", "text": user_prompt}]},
    ]

    inputs = processor.apply_chat_template(
        messages,
        tokenize=True,
        add_generation_prompt=True,
        return_tensors="pt",
        return_dict=True
    )

    inputs = {k: v.to(model.device) for k, v in inputs.items()}

    output_ids = model.generate(
        **inputs,
        **gen_kwargs,
        eos_token_id=processor.tokenizer.eos_token_id
    )

    generated_ids = output_ids[0][inputs["input_ids"].shape[-1]:]
    return processor.tokenizer.decode(generated_ids, skip_special_tokens=True).strip()


# ============================================================
# MAIN FUNCTION
# ============================================================

def generate_excel_from_multiple_txt(template_xlsx_path,
                                     qwen_txt_paths,
                                     model,
                                     processor,
                                     gen_kwargs,
                                     url_list):

    instrucciones_path = "instrucciones.txt"

    # 1. Leer instrucciones.txt
    with open(instrucciones_path, "r", encoding="utf-8") as f:
        instrucciones = f.read()

    # 2. Abrir plantilla
    wb = load_workbook(template_xlsx_path)
    ws = wb.worksheets[0]

    # 3. Leer encabezados de la plantilla (fila 4)
    headers = [ws.cell(row=4, column=i).value for i in range(1, 15)]

    current_row = 5  # fila donde empiezan los datos

    # 4. Procesar cada QWEN.txt
    for qwen_path, url in zip(qwen_txt_paths, url_list):

        with open(qwen_path, "r", encoding="utf-8") as f:
            qwen_text = f.read()

        # 5. Construir prompt para Qwen
        system_prompt = (
            "Eres un motor de extracción de datos.\n"
            "Debes rellenar EXACTAMENTE las columnas del Excel siguiendo los encabezados proporcionados.\n"
            "Devuelve SOLO filas con columnas separadas por |||.\n"
            "No añadas explicaciones.\n"
            "No añadas texto antes o después.\n"
            "No uses JSON.\n"
            "No uses markdown.\n"
        )

        user_prompt = (
            "Rellena las 14 columnas del Excel siguiendo estas instrucciones:\n\n"
            f"{instrucciones}\n\n"
            "Usa ÚNICAMENTE la información del siguiente texto:\n\n"
            f"{qwen_text}\n\n"
            "Devuelve SOLO las 14 columnas separadas por |||.\n"
            "Ejemplo de formato:\n"
            "col1 ||| col2 ||| col3 ||| ... ||| col14\n"
        )


        # 6. Llamar a Qwen
        raw_output = qwen_generate(model, processor, system_prompt, user_prompt, gen_kwargs)

        print("\n=== RAW OUTPUT FROM QWEN ===")
        print(raw_output)
        print("=== END RAW OUTPUT ===\n")



        # 7. Extraer solo líneas válidas
        valid_lines = [line for line in raw_output.splitlines() if "|||" in line]

        # 8. Escribir cada fila en el Excel
        for line in valid_lines:
            cols = [c.strip() for c in line.split("|||")]

            # Ajustar número de columnas
            if len(cols) < len(headers):
                cols += [""] * (len(headers) - len(cols))
            elif len(cols) > len(headers):
                cols = cols[:len(headers)]

            # Escribir en Excel
            for j, value in enumerate(cols):
                ws.cell(row=current_row, column=1 + j).value = value

            current_row += 1

    # 9. Guardar Excel final
    output_path = "output_final.xlsx"
    wb.save(output_path)
    return output_path

# import os
# from openpyxl import load_workbook
# import torch

# # ============================================================
# # QWEN GENERATION
# # ============================================================

# @torch.inference_mode()
# def qwen_generate(model, processor, system_prompt, user_prompt, gen_kwargs):
#     messages = [
#         {"role": "system", "content": [{"type": "text", "text": system_prompt}]},
#         {"role": "user", "content": [{"type": "text", "text": user_prompt}]},
#     ]

#     inputs = processor.apply_chat_template(
#     messages,
#     tokenize=True,
#     add_generation_prompt=True,
#     return_tensors="pt",
#     return_dict=True
# )

#     inputs = {k: v.to(model.device) for k, v in inputs.items()}

#     output_ids = model.generate(
#         **inputs,
#         **gen_kwargs,
#         eos_token_id=processor.tokenizer.eos_token_id
#     )

#     generated_ids = output_ids[0][inputs["input_ids"].shape[-1]:]
#     return processor.tokenizer.decode(generated_ids, skip_special_tokens=True).strip()


# # ============================================================
# # MAIN FUNCTION
# # ============================================================

# def generate_excel_from_multiple_txt(template_xlsx_path,
#                                      qwen_txt_paths,
#                                      model,
#                                      processor,
#                                      gen_kwargs,
#                                      url_list
#                                      ):

#     instrucciones_path="instrucciones.txt"
#     # 1. Leer instrucciones.txt
#     with open(instrucciones_path, "r", encoding="utf-8") as f:
#         instrucciones = f.read()

#     # 2. Abrir plantilla
#     wb = load_workbook(template_xlsx_path)
#     ws = wb.worksheets[0]

#     current_row = 5  # fila donde empiezan los datos

#     # 3. Procesar cada QWEN.txt
#     for qwen_path, url in zip(qwen_txt_paths, url_list):

#         with open(qwen_path, "r", encoding="utf-8") as f:
#             qwen_text = f.read()

#         # 4. Construir prompt para Qwen
#         system_prompt = (
#             "Eres un motor de extracción de datos.\n"
#             "Debes rellenar EXACTAMENTE 14 columnas por receta, siguiendo las instrucciones proporcionadas.\n"
#             "Devuelve SOLO filas con 14 columnas separadas por |||.\n"
#             "No añadas explicaciones.\n"
#             "No añadas texto antes o después.\n"
#             "No uses JSON.\n"
#             "No uses markdown.\n"
#         )

#         user_prompt = (
#             f"INSTRUCCIONES:\n{instrucciones}\n\n"
#             f"URL DEL VIDEO: {url}\n\n"
#             f"TEXTO A PROCESAR:\n{qwen_text}\n\n"
#             "Devuelve una fila por cada receta encontrada.\n"
#             "Cada fila debe tener EXACTAMENTE 14 columnas separadas por |||."
#         )

#         # 5. Llamar a Qwen
#         raw_output = qwen_generate(model, processor, system_prompt, user_prompt, gen_kwargs)

#         # 6. Procesar cada fila devuelta por Qwen
#         for line in raw_output.splitlines():
#             if "|||" not in line:
#                 continue  # ignorar líneas incorrectas

#             cols = [c.strip() for c in line.split("|||")]

#             # Asegurar 14 columnas
#             if len(cols) < 14:
#                 cols += [""] * (14 - len(cols))
#             elif len(cols) > 14:
#                 cols = cols[:14]

#             # 7. Escribir en Excel
#             for j, value in enumerate(cols):
#                 ws.cell(row=current_row, column=1 + j).value = value

#             current_row += 1

#     # 8. Guardar Excel final
#     output_path = "output_final.xlsx"
#     wb.save(output_path)
#     return output_path

