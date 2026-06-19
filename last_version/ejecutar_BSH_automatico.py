import os
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import json
from tkinterdnd2 import DND_FILES, TkinterDnD
import csv
import re
import pandas as pd
import string
from openpyxl import load_workbook
import aux_fun as aux
import sys
from openpyxl import load_workbook

# --- Common English stopwords ---
STOPWORDS = {
    "a", "an", "and", "the", "of", "in", "on", "at", "to", "for", "by", "with",
    "from", "about", "as", "into", "like", "through", "after", "over", "between",
    "out", "against", "during", "without", "before", "under", "around", "among",
    # Español
    "de", "la", "el", "y", "en", "con", "para", "por", "del", "los", "las", "un", "una"
}


# ============================================================
# VARIABLES
# ============================================================

# CASE 0 (0 total matches)
ID_not_found = 'IDNotFound'

# CASE 1 or CASE 2 (1 exact or partial match)
#   Sí es cooking method y coincide --> se rellena el ID 
#   XXXX
#   Sí es cooking method pero NO coincide (el mismo para CASE 3 or CASE 4)
ID_yes_cooking_method_but_not_match = 'IDFound.CookingMethod'

#   No es cooking method (el mismo para CASE 3 or CASE 4)
ID_no_cooking_method = 'IDFound.NoCookingMethod'

#   El cooking method es NEW (el mismo para CASE 3 or CASE 4)
ID_new_cooking_method = 'IDFound.NEWPreparationMethod'


# CASE 3 or CASE 4 (+1 exact or partial match)
#   Sí es cooking method y ninguno coincide (ID_yes_cooking_method_but_not_match)

#   Sí es cooking method y coincide con uno solo --> se rellena el ID
#   XXXX

#   Sí es cooking method y coinciden varios y determinamos cuál de todos es --> se rellena el ID
#   XXXX

#   Sí es cooking method y coinciden varios y no determinamos cuál de todos es

#   No es un Cooking Method -> ID_no_cooking_method

#   El cookign method es NEW -> ID_new_cooking_method




# ============================================================
# OBTENER ARCHIVOS SIN TKINTER
# ============================================================

# Archivo 1: fijo, siempre el mismo
archivo1 = "ID-Dish-Catalogue_v2.xlsx"
archivo1_nombre = os.path.basename(archivo1)

# Archivo 2: viene desde main.py como argumento
if len(sys.argv) < 2:
    raise ValueError("No se ha proporcionado el archivo Excel generado por IA.")

archivo2 = sys.argv[1]
archivo2_nombre = os.path.basename(archivo2)

print("Archivo base de datos:", archivo1)
print("Archivo generado por IA:", archivo2)


# ============================================================
# MAIN CODE
# ============================================================
# Main function to execute the program
#  * file 1 = ID-Dish-Catalogue_v2.xlsx
#  * file 2 = colores.xlsx
def program():

    file1 = archivo1_nombre
    file2 = archivo2_nombre

    print('Ejecutando el código...\n\n')

    # --- Step 1: Convert Excels files to CSVs ---
    file_dataBase_excel = file1
    file_dataBase_csv = file1.replace('.xlsx', '.csv')
    aux.from_excel_to_csv(file_dataBase_excel, file_dataBase_csv)
    file_colores_excel = file2
    file_colores_csv = file2.replace('.xlsx', '.csv')
    aux.from_excel_to_csv(file_colores_excel, file_colores_csv)

    # --- Step 2: Load the CSV files into DataFrames and obtain the KPI values and indexes ---
    df_colores = pd.read_csv(file_colores_csv, sep=';', dtype=str)
    df_base = pd.read_csv(file_dataBase_csv, sep=';', dtype=str)

    # Get the KPI values from the third row (KPI values) → save as `list of tuples` (index, weight)
    index_numerico = aux.get_kpi_values_and_indexes(file_colores_csv)









    # --- Step 3: Sort the KPI values in descending order to prioritize the most important KPIs in the search ---
    index_numerico.sort(key=lambda x: x[1], reverse=True)

    # Get the first index and value of the KPI to use as the main KPI for the search (the one with the highest weight).
    index_KPI = 0
    actual_value_KPI = index_numerico[index_KPI][1]
    
    # Print the first row of each CSV file to check the column names and save them in an array
    array_head_colores, array_head_base = aux.print_first_row_of_csv(file_colores_csv, file_dataBase_csv)

    # --- Step 4: Preprocess the base Excel file ---
    # Remove unnecessary rows from df_colores
    df_colores = df_colores.iloc[3:]


    columnas_archivo1 = [
        "contentData.name", "contentData.dishType", "contentData.imageResource",
        "contentData.applianceType", "contentData.technicalDescription",
        "contentData.deprecated", "contentData.properties.preparationType",
        "contentData.properties.category", "contentData.properties.vegetableFreshOrFrozen",
        "contentData.properties.vegetableBreaded", "contentData.properties.vegetablePreparationMethod",
        "contentData.temperatureLevel", "contentData.fryingSensorTemperature",
        "contentData.durationRangeInSeconds.minimum", "contentData.durationRangeInSeconds.maximum",
        "id", "contentData.properties.meatBone", "contentData.properties.meatDimension",
        "contentData.properties.meatOutcome", "contentData.properties.meatPreparationMethod",
        "contentData.properties.pastaRiceAndLegumesDriedOrFresh", "contentData.properties.pastaRiceAndLegumesPreparationMethod",
        "contentData.properties.pastaRiceAndLegumesGrainType", "contentData.properties.friedFoodCookingFat",
        "contentData.properties.friedFoodPreparationMethod", "contentData.properties.vegetableOutcome",
        "contentData.properties.meltingPreparationMethod", "contentData.properties.fishCut",
        "contentData.properties.fishFreshOrFrozen", "contentData.properties.fishPreparationMethod",
        "contentData.properties.fishBreaded", "contentData.properties.friedFoodFreshOrFrozen",
        "contentData.properties.friedFoodBreaded", "contentData.technicalImages.src",
        "contentData.technicalImages.title", "contentData.properties.poultryPreparationMethod",
        "contentData.properties.poultryBreaded", "contentData.properties.vegetablePrecookedOrRaw",
        "contentData.properties.meltingChocolate", "contentData.properties.friedFoodThickness",
        "contentData.properties.meatBreaded", "contentData.properties.meatPrecookedOrRaw",
        "contentData.properties.meatType", "contentData.properties.meatDoneness",
        "contentData.properties.meatThickness", "contentData.properties.meatRind",
        "contentData.properties.friedFoodMainIngredient"
    ]

    columnas_archivo2 = [
        'ID-ord', 'LINK', 'Recipe Title', 'Recipe ingredients', 'OPTIONAL', 'CookwareSession',
        'Step-ordinal', 'Ingredients', 'Added ingredients', 'Main ingredient', 'Cooking Process',
        'Outcome', 'Duration', 'LID', 'ID-Dish OUTPUT', 'ID-Dish OUTPUT.1', 'KPI WEIGHT', 'DRAFT REQUIRED POWER', 'MAIN INGREDIENT', 'CATEGORY MAIN INGREDIENT', 'Propoerties', 'Property1: FreshOrFrozen', 'Property 2: PrecookedOrRaw', 'Property 3: PreparationMethod', 'Property 4: Breaded', 'Property 5: Cut', 'Property 6: Outcome (crispy/soft)', 'Property 7: Doneness', 'Property 8: meatBone', 'Property 9: meatThickness', 'Property 10:  meatDimension', 'Property 11: meatType', 'Property 12: pastaRiceAndLegumesDriedOrFresh', 'Property 13: pastaRiceAndLegumesGrainType', 'Property 14: friedFoodThickness', 'Property 15: friedFoodCookingFat'
    ]


    # Save the column mapping in a JSON file: mapeo_columnas.json
    aux.save_column_mapping_to_json(columnas_archivo1, columnas_archivo2, array_head_base, array_head_colores)
    name_file = "mapeo_columnas.json"
    index_database_1, index_colores_2 = aux.obtener_valores_json(name_file)


    # Get the column index of:
    index_main_ingredient_colores_2 = aux.get_column_index(index_colores_2[9])
    index_cooking_method_colores_2 = aux.get_column_index(index_colores_2[10]) # actual columna K
    index_columna_L_colores_2 = aux.get_column_index(index_colores_2[11])
    index_columna_AA_colores_2 = aux.get_column_index(index_colores_2[26])
    index_columna_H_colores_2 = aux.get_column_index(index_colores_2[7])

    index_columna_I_colores_2 = aux.get_column_index(index_colores_2[8])
    index_columna_O_colores_2 = aux.get_column_index(index_colores_2[14])
    index_columna_P_colores_2 = aux.get_column_index(index_colores_2[15])
    index_columna_Q_colores_2 = aux.get_column_index(index_colores_2[16])
    index_columna_R_colores_2 = aux.get_column_index(index_colores_2[17])
    index_columna_S_colores_2 = aux.get_column_index(index_colores_2[18])
    index_columna_T_colores_2 = aux.get_column_index(index_colores_2[19])

    index_columna_V_colores_2 = aux.get_column_index(index_colores_2[21])

    index_columna_X_colores_2 = aux.get_column_index(index_colores_2[23])

    index_lenght_colores_2 = len(index_colores_2) -1

    index_ID_database_1 = aux.get_column_index(index_database_1[15])
    contentData_properties_friedFoodMainIngredient = aux.get_column_index(index_database_1[46])

    #############################################
    #                 MAIN LOOP                 #
    #############################################

    # Set the initial value of the KPI sum to 0 for each row
    sumador_KPIS_valor = 0
    sumatorio_KPIS = 0
    index_KPI = 0

    # Loop through each row in the df_colores DataFrame
    for index, row in df_colores.iterrows():


        sumatorio_KPIS, index_KPI, new_ID, new_ID_cooking_method_ = 0, 0, '', ''
        value_case_cooking_method = 3  # Initialize the value of the cooking method case (3 = error, 0 = new, 1 = cooking method, 2 = no cooking method)
        actual_value_KPI = index_numerico[index_KPI][1]

        # Get the main ingredient 
        main_ingredient_to_search = str(row.iloc[index_main_ingredient_colores_2])
        print(f"\nProcessing row {index} - Main ingredient to search: '{main_ingredient_to_search}'")
        print (f"Cooking method to search: '{row.iloc[index_cooking_method_colores_2]}'")

        case = 0    # Initialize CASE variable 
                    # 0 = NO MATCHES
                    # 1 = EXACT
                    # 2 = PARTIAL MATCH
                    # 3 = >1 MATCHES EXACTOS
                    # 4 = >1 MATCHES PARCIALES

        # Get the number of matches 
        case, exact_matches, partial_matches = aux.get_number_of_matches(case, df_base, main_ingredient_to_search, STOPWORDS)

        if case == 0:
            print('-')
            print('   -----------------------0 matches (totales ni parciales)')

            new_ID = ID_not_found
            new_ID_cooking_method_ = aux.concatenate_with_dot(main_ingredient_to_search, row.iloc[index_cooking_method_colores_2])
            sumador_KPIS_valor = 0

            print(f' --> new_ID: {new_ID}')
            print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
            print(f' --> Valor KPI: {sumador_KPIS_valor}')

        elif case == 1 or case == 2:
            print('-')
            print('   -----------------------1 match (exacto o parcial)')

            # Get the match        
            matches = exact_matches if case == 1 else partial_matches
            
            # Check if the cooking method is a cooking method, if it is not, or if it is new
            #   --> Case 0: It's NEW
            #   --> Case 1: It's a cooking method
            #   --> Case 2: It's NOT a cooking method
            #   --> Case 3: An error has ocurred

            value_case_cooking_method = aux.check_cooking_method(row, index_cooking_method_colores_2, file_colores_excel)

            index_KPI = index_KPI + 1
            actual_value_KPI = index_numerico[index_KPI][1] + actual_value_KPI

            sumador_KPIS_valor = actual_value_KPI

            # print("Lo dejamos preparado en el siguiente index_KPI para el siguiente paso, por si queremos seguir sumando KPIs:")
            index_KPI = index_KPI + 1

            print("KPI ACTUAL, ", sumador_KPIS_valor)



            if value_case_cooking_method == 0: # IT IS NEW
                
                # new_ID = 'IDFound.NEWPreparationMethod'
                new_ID = ID_new_cooking_method

                # The cooking method is NEW (it isn't in the list of cooking methods)
                # Example: main ingredient = "King Prawns", cooking method = "NEW: air fried"

                # Get tht ingredient from the matches (it can be the exact match or the partial match, it depends on the case) and save it in a variable
                ingredient = matches.iloc[0, 0]

                # Get the value of the cooking method column ("NEW: air fried")
                cooking_method = row.iloc[index_cooking_method_colores_2]
 
                # Delete "new:" and the spaces at the beginning and the end ("air fried")
                cooking_method = cooking_method.replace("NEW:", "").strip()

                # Concatenate the main ingredient and the cooking method with a dot in between, and convert it to lowercase (MainIngredient.AirFried)
                new_ID_cooking_method_ = aux.concatenate_with_dot(ingredient, cooking_method)

              
            elif value_case_cooking_method == 1: # IT IS A COOKING METHOD

                # Check if the cooking method matches with the cooking method of the ingredient in the catalogue
                
                # Get the value of the cooking method column ("Fry")
                cooking_method = row.iloc[index_cooking_method_colores_2]

                # Convert the cooking method in minus
                cooking_method = cooking_method.lower() 

                # Delete all the spaces in the cooking method
                cooking_method = cooking_method.replace(" ", "")

                # Get the cooking method format for the catalogue
                cooking_method_catalogue_format = aux.convert_cooking_method_format(cooking_method)

                # Check if the cooking method converted matches (exists) in the match (catalogue)
                subset = matches.iloc[:, 9:]

                coincidences = subset == cooking_method_catalogue_format

                hay_coincidencia = False

                if coincidences.any().any():
                    hay_coincidencia = True
                    print("Hay coincidencia")

                    coincidentes = coincidences.stack()
                    coincidentes = coincidentes[coincidentes]

                    for fila_idx, col_nombre in coincidentes.index:
                        valor = matches.loc[fila_idx, col_nombre]
                        print(f"Coincide en fila {fila_idx}, columna '{col_nombre}': {valor}")

                        id_temporal_2 = matches.loc[fila_idx].iloc[0]
                        print(f"Valor de la columna 0 de esa fila: {id_temporal_2}")

                else:
                    print("NO HAY COINCIDENCIA")

               
                if hay_coincidencia: # The cooking method matches
                    print("-------------------PRUEBA DE QUE HA COINCIDIDO EL COOKING PROCESS")
                    
                    cooking_process_catalogue =  matches.loc[fila_idx, col_nombre]
                    ingredient_catalogue = matches.loc[fila_idx].iloc[0]
                    id_catalogue = matches.loc[fila_idx].iloc[15]
                    print(cooking_process_catalogue)
                    print(ingredient_catalogue)
                    print(id_catalogue)


                    # KingPrawns -> Kingprawns (only the first letter in cap)
                    # Format de ingredient to have the first letter of each word in uppercase and the rest in lowercase ("KingPrawns -> Kingprawns")
                    ingredient = string.capwords(ingredient_catalogue)

                    # Format de cooking method to have the first letter of each word in uppercase and the rest in lowercase ("cut" -> Cut)
                    cooking_method = string.capwords(cooking_method)


                    new_ID = id_catalogue
                    new_ID_cooking_method_ = ingredient + '.' + cooking_method
                    # sumador_KPIS_valor = actual_value_KPI
                    print(f' --> new_ID: {new_ID}')
                    print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
                    print(f' --> Valor KPI: {sumador_KPIS_valor}')
                
                else: # the cooking methods doesnt match


                    cooking_process_catalogue =  matches.loc[fila_idx, col_nombre]
                    ingredient_catalogue = matches.loc[fila_idx].iloc[0]
                    id_catalogue = matches.loc[fila_idx].iloc[15]
                    print(cooking_process_catalogue)
                    print(ingredient_catalogue)
                    print(id_catalogue)

                    # KingPrawns -> Kingprawns (only the first letter in cap)
                    # Format de ingredient to have the first letter of each word in uppercase and the rest in lowercase ("KingPrawns -> Kingprawns")
                    ingredient = string.capwords(ingredient_catalogue)

                    # Format de cooking method to have the first letter of each word in uppercase and the rest in lowercase ("cut" -> Cut)
                    cooking_method = string.capwords(cooking_method)

                    new_ID = ID_yes_cooking_method_but_not_match 
                    new_ID_cooking_method_ = ingredient + '.' + cooking_method

                

            elif value_case_cooking_method == 2: # IT IS NOT A COOKING METHOD

                new_ID = ID_no_cooking_method 

                # The cooking method is NOT a cooking method, so we have to concatenate the main ingredient with that no Cooking Method
                # Example: the ingredient from the catalogue that matched is KingPrawns and the cooking method is Cut --> "Kingprawns.Cut"

                # Get the value of the cooking method column ("Cut")
                cooking_method = row.iloc[index_cooking_method_colores_2]

                # Get tht ingredient from the matches (it can be the exact match or the partial match, it depends on the case) and save it in a variable
                ingredient = matches.iloc[0, 0]

                new_ID_cooking_method_ = aux.concatenate_with_dot(ingredient, cooking_method)

                sumador_KPIS_valor = actual_value_KPI
                
            print(f' --> new_ID: {new_ID}')
            print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
            print(f' --> Valor KPI: {sumador_KPIS_valor}')
                

        elif case == 3 or case == 4: # MORE THAN ONE MATCHES (EXACT OR PARTIAL)
            print('   -----------------------+ 1 match (exacto o parcial)')


            # print("Como estamos mirando el proceso de cocinado, aumentamos el KPI")

            # # Actualmente KPI = (main ingredient) y queremos aumentar el KPI cono + peso del cooking process
            # print("KPI ACTUAL, ", actual_value_KPI) # 70
            # print("index KPI ACTUAL, ", index_KPI)  # 0

            # print ( "AUMENTAMOS PESO KPI")

            index_KPI = index_KPI + 1
            actual_value_KPI = index_numerico[index_KPI][1] + actual_value_KPI

            sumador_KPIS_valor = actual_value_KPI

            # print("Lo dejamos preparado en el siguiente index_KPI para el siguiente paso, por si queremos seguir sumando KPIs:")
            index_KPI = index_KPI + 1

            # print("KPI ACTUAL, ", sumador_KPIS_valor)


            # Get the matches      
            more_than_one_match = exact_matches if case == 3 else partial_matches
            
            # Check if the cooking method is a cooking method, if it is not, or if it is new
            #   --> Case 0: It's NEW
            #   --> Case 1: It's a cooking method
            #   --> Case 2: It's NOT a cooking method
            #   --> Case 3: An error has ocurred

            value_case_cooking_method = aux.check_cooking_method(row, index_cooking_method_colores_2, file_colores_excel)
            print(f"Value case cooking method: {value_case_cooking_method}")    

            if value_case_cooking_method == 0: # IT IS NEW

                new_ID = ID_new_cooking_method

                # Get the cooking method and main ingredient 
                # Example: "NEW: air fried" and "King Prawns"
                cooking_method = row.iloc[index_cooking_method_colores_2]
                ingredient = main_ingredient_to_search

                # Remove NEW
                cooking_method = cooking_method.replace("NEW:", "").strip()

                # Format the words and obtains: "KingPrawns.AirFried"
                new_ID_cooking_method_ = aux.concatenate_with_dot(ingredient, cooking_method)

                sumador_KPIS_valor = actual_value_KPI

                print(f' --> new_ID: {new_ID}')
                print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
                print(f' --> Valor KPI: {sumador_KPIS_valor}')

                
            elif value_case_cooking_method == 1: # IT IS A COOKING METHOD


                # Check if the cooking method matches with the cooking method of the any of the ingredients that we matched on the catalogue

                # Get the value of the cooking method column color excel ("Fry")
                cooking_method = row.iloc[index_cooking_method_colores_2]

                # Convert the cooking method in minus
                cooking_method = cooking_method.lower() 

                # Delete all the spaces in the cooking method
                cooking_method = cooking_method.replace(" ", "")

                # Get the cooking method format for the catalogue
                cooking_method_catalogue_format = aux.convert_cooking_method_format(cooking_method)
                print(f"Cooking method in the format of the catalogue: '{cooking_method_catalogue_format}'")

                # Check if the cooking method converted matches (exists) in the match (catalogue)
                subset = more_than_one_match.iloc[:, 9:]
        
                coincidences = subset == cooking_method_catalogue_format

                print("Coincidences between the cooking method and the catalogue:")
                print(coincidences)

                hay_coincidencia = 0 # 0 no hay, 1, hay 1, 2, hay mas de 1

                filas_coincidentes = {}

                print('¿Cuantas coincidencias?')
                if coincidences.any().any():
                    coincidentes = coincidences.stack()
                    coincidentes = coincidentes[coincidentes]

                    filas_coincidentes = set(fila_idx for fila_idx, col_nombre in coincidentes.index)

                    print("Filas con coincidencia:", filas_coincidentes)
                    print("Número de filas distintas con coincidencia:", len(filas_coincidentes))

                    if len(filas_coincidentes) == 1:
                        print("solo ha coincidido un total de una vez")

                        hay_coincidencia = 1
                    else:
                        print("rellenar para decidir")

                        hay_coincidencia = 2
                else:
                    print("no coincidencias")

                    hay_coincidencia = 0
                    

                if hay_coincidencia == 0: # El cooking method no coincide con ningún cooking method de las coincidencias enocntradas

                    new_ID = ID_yes_cooking_method_but_not_match
                    
                    # Get the value of the cooking method column ("Cut")
                    cooking_method = row.iloc[index_cooking_method_colores_2]

                    # Get the main ingredient and save it in a variable
                    ingredient = main_ingredient_to_search

                    new_ID_cooking_method_ = aux.concatenate_with_dot(ingredient, cooking_method)

                    sumador_KPIS_valor = actual_value_KPI

                    print(f' --> new_ID: {new_ID}')
                    print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
                    print(f' --> Valor KPI: {sumador_KPIS_valor}')
        

                elif hay_coincidencia == 1: # El cooking method coincide con un único ingrediente (solo coincide con un proceso de cocinado)

                    # Tomamos la primera coincidencia
                    fila_idx, col_nombre = coincidentes.index[0]

                    print(f"El cooking method coincide con la fila {fila_idx} y la columna {col_nombre} del catálogo.")

                    # cooking_process_catalogue = matches.loc[fila_idx, col_nombre]
                    # ingredient_catalogue = matches.loc[fila_idx].iloc[0]
                    # id_catalogue = matches.loc[fila_idx].iloc[15]


                    fila_real = more_than_one_match.iloc[-1]   # última fila, la de datos reales
                    cooking_process_catalogue = fila_real[col_nombre]

                    ingredient_catalogue = fila_real[0]

                    id_catalogue = fila_real[15]

                    # Get the value of the cooking method column ("Cut")
                    cooking_method = row.iloc[index_cooking_method_colores_2]

                    # Get the main ingredient and save it in a variable
                    ingredient = main_ingredient_to_search

                    new_ID_cooking_method_ = aux.concatenate_with_dot(ingredient, cooking_method)

                    new_ID = id_catalogue
                    sumador_KPIS_valor = actual_value_KPI

                    print(f' --> new_ID: {new_ID}')
                    print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
                    print(f' --> Valor KPI: {sumador_KPIS_valor}')

                elif hay_coincidencia == 2: # El cooking method coincide con varios procesos de cocinado

                    print("RELLENAMOS COLUMNAS AZULES PARA PODER DECIDIR")
                    # Rellenamos las columnas azules del excel de colores de esa fila que estamos mirando
                    # Una vez rellenada la recorremos 

                    column_J_colores = main_ingredient_to_search # no lo modificamos
                    column_K_colores = row.iloc[index_cooking_method_colores_2]
                    column_L_colores = row.iloc[index_columna_L_colores_2]

                    print(f'column_J_colores: {column_J_colores}')
                    print(f'column_K_colores: {column_K_colores}')
                    print(f'column_L_colores: {column_L_colores}')

                    column_J_colores = main_ingredient_to_search # no lo modificamos

                    # Convertimos el proceso de cocinado al formato que coincida con el catálogo de ingredientes
                    column_K_colores = row.iloc[index_cooking_method_colores_2]
                    column_K_colores = column_K_colores.lower() # Convert the cooking method in minus
                    column_K_colores = column_K_colores.replace(" ", "") # Delete all the spaces in the cooking method
                    column_K_colores  = aux.convert_cooking_method_format(column_K_colores) # Get the cooking method format for the catalogue

                    # Obtenemos el outcome y lo convertimos a mayusculas todo
                    column_L_colores = row.iloc[index_columna_L_colores_2]
                    column_L_colores = column_L_colores.upper()
                    #eliminamos espacios en blanco
                    column_L_colores = column_L_colores.replace(" ", "")


                    print(f'column_J_colores: {column_J_colores}')
                    print(f'column_K_colores: {column_K_colores}')
                    print(f'column_L_colores: {column_L_colores}')

                    column_S_colores = column_J_colores
                    column_X_colores = column_K_colores
                    column_AA_colores = column_L_colores

                    # Guardamos los valores en las columnas azules
                    df_colores.at[index, df_colores.columns[index_columna_S_colores_2]] = column_S_colores
                    df_colores.at[index, df_colores.columns[index_columna_X_colores_2]] = column_X_colores
                    df_colores.at[index, df_colores.columns[index_columna_AA_colores_2]] = column_AA_colores

                    # comprobamos si el valor de L, ahora AA coincide con algo de las coincidencias encontradas
                    # Recorremos las columnas del excel de colores en orden decreciente del peso de los KPIs



                    print("REVISAR DESDE AQUÍ----")

                    print("index KPI, ", index_KPI)
                    print("KPI ACTUAL, ", actual_value_KPI)

                    # index_KPI = index_KPI + 1
                    # actual_value_KPI = index_numerico[index_KPI][1]

                    # print("---- ---- : miramos el siguiente valor del KPI:")
                    # print("index KPI, ", index_KPI)
                    # print("KPI ACTUAL, ", actual_value_KPI)

                    # sumador_KPIS_valor = actual_value_KPI + sumador_KPIS_valor
                    # print("sumamos: ", sumador_KPIS_valor)

                    print("------>")
                    candidatas = subset.loc[list(filas_coincidentes)]
                    print ("candidatas: ", candidatas)

                    # obtener la columan A del catalogo de ingrediente de las filas candidasas
                    print("Columna A del catálogo de ingredientes de las filas candidatas:")
                    for idx, fila in candidatas.iterrows():
                        id_temporal_2 = more_than_one_match.loc[idx].iloc[0]
                        print(f"Fila {idx}: {id_temporal_2}, {fila.iloc[1]}")



                    encontrada = False
                    sumatorio = sumador_KPIS_valor
                    print("sumatorio antes empezar: ", sumatorio)

                    pares = index_numerico
                    # print("pares: ", pares)

                    
                    # print(pares[2:])
                    
                    for idx, valor_sumar in pares[2:]:   # empezamos en el tercer par

                        sumatorio += valor_sumar

                        valor_a_comprobar = row.iloc[idx]
                        print(f"\nComprobando valor del KPI en columna index {idx} con valor '{valor_a_comprobar}' y sumatorio actual {sumatorio}")

                        coincidences_valor = candidatas == valor_a_comprobar
                        print("coincidences_valor: ", coincidences_valor)

                        coincidentes_valor = coincidences_valor.stack()
                        coincidentes_valor = coincidentes_valor[coincidentes_valor]

                        filas_filtradas = set(fila_idx for fila_idx, col_nombre in coincidentes_valor.index)

                        print("Filas que contienen el valor:", filas_filtradas)

                        if len(filas_filtradas) == 1:
                            print("ENCONTRADO:", list(filas_filtradas)[0])
                            encontrada = True

                            f_i, col_nombre = coincidentes_valor.index[0]
                            id_catalogue = more_than_one_match.loc[f_i].iloc[15]

                            
                            cooking_method = row.iloc[index_cooking_method_colores_2]   # Get the value of the cooking method column ("Cut")   
                            ingredient = main_ingredient_to_search  # Get the main ingredient and save it in a variable

                            # Concatenate the main ingredient and the cooking method with a dot in between, and convert it to lowercase (MainIngredient.AirFried)
                            new_ID_cooking_method_ = aux.concatenate_with_dot(ingredient, cooking_method)

                            new_ID = id_catalogue

                            break

                        elif len(filas_filtradas) > 1:
                            candidatas = candidatas.loc[list(filas_filtradas)]



                    if not encontrada:
                        print("MULTIPLES OPCIONES")
                        # print(candidatas)


                        # Recuperar las filas candidatas completas, con todas las columnas originales
                        candidatas_completas = more_than_one_match.loc[candidatas.index]

                        # Valor de la primera coincidencia, columna 15 por POSICIÓN
                        valor = candidatas_completas.iloc[0, 15]

                        # print("Valor de la primera coincidencia, columna 15:", valor)
                        new_ID = valor




        

                        


                        # Get the main ingredient and save it in a variable
                        ingredient = main_ingredient_to_search
                        ingredient = string.capwords(ingredient)
                        # Format de ingredient to have the first letter of each word in uppercase and the rest in lowercase ("KingPrawns -> Kingprawns")
                        ingredient = string.capwords(ingredient)
                        # Delete all the spaces in the ingredient ("Ejemplo Prueba -> EjemploPrueba")
                        ingredient = ingredient.replace(" ", "")

                        cooking_method = row.iloc[index_cooking_method_colores_2]
                        cooking_method = cooking_method.lower() 
                        cooking_method = cooking_method.replace(" ", "")
                        cooking_method = string.capwords(cooking_method)


                        lineas = new_ID_cooking_method_


                        for _, fila in candidatas_completas.iterrows():
                            val15 = fila.iloc[15]
                            val0 = fila.iloc[0]
                            # lineas += f"({val15}, {val0})\n"
                            lineas += f"{val0}.{cooking_method}, {val15}\n"



                        new_ID_cooking_method_ = ingredient + '.' + "MultipleOptions" + '\n' + lineas





                    sumador_KPIS_valor = sumatorio
                    print("sumatorio finalizar: ", sumador_KPIS_valor)






            elif value_case_cooking_method == 2: # IT IS NOT A COOKING METHOD

                new_ID = ID_no_cooking_method

                # NOT A COOKING METHOD: Concatenate main ingredient the no cooking method ("King Prawns" and "Cut" --> "KingPrawns.Cut"

                # Get the value of the cooking method column ("Cut")
                cooking_method = row.iloc[index_cooking_method_colores_2]

                # Get the main ingredient and save it in a variable ("King Prawns")
                main_ingredient = main_ingredient_to_search

                new_ID_cooking_method_ = aux.concatenate_with_dot(main_ingredient, cooking_method)
                sumador_KPIS_valor = actual_value_KPI

                print(f' --> new_ID: {new_ID}')
                print(f' --> new_ID_cooking_method_: {new_ID_cooking_method_}')
                print(f' --> Valor KPI: {sumador_KPIS_valor}')





        df_colores.at[index, df_colores.columns[index_columna_O_colores_2]] = new_ID
        df_colores.at[index, df_colores.columns[index_columna_P_colores_2]] = new_ID_cooking_method_
        df_colores.at[index, df_colores.columns[index_columna_Q_colores_2]] = sumador_KPIS_valor

    df_colores.to_csv(file_colores_csv, sep=';', index=False, encoding='utf-8')

    file_name = aux.obtain_file_name_1(file2)

    wb = load_workbook(file2)
    # ws = wb.active
    ws = wb.worksheets[0]

    # Define los índices de inicio y fin de las columnas que quieres procesar
    indice_inicio = index_columna_O_colores_2
    indice_fin = index_lenght_colores_2

    # Define la columna de inicio en la hoja de cálculo donde quieres empezar a escribir
    columna_excel_inicio = 15  # Por ejemplo, la columna O

    # Itera a través del rango de columnas del DataFrame
    for i, indice_columna in enumerate(range(indice_inicio, indice_fin + 1)):
        # Extrae la columna actual como una lista
        columna_lista = df_colores.iloc[:, indice_columna].tolist()

        # Calcula el número de columna en la hoja de cálculo (ahora de forma consecutiva)
        columna_excel = columna_excel_inicio + (indice_columna - indice_inicio)

        # Escribe los valores de la lista en la columna correspondiente de la hoja de cálculo
        for j, valor in enumerate(columna_lista, start=5):
            ws.cell(row=j, column=columna_excel, value=valor)


    # Si ya has realizado otras modificaciones en ws, guarda el libro de trabajo
    # wb.save("nombre_de_tu_archivo.xlsx")


    wb.save(file_name + "_actualizado.xlsx")
    wb.close()
    # borrar ficher json
    if os.path.exists("mapeo_columnas.json"):
        os.remove("mapeo_columnas.json")
        print(f"El fichero mapeo_columnas.json ha sido borrado.")
    else:
        print(f"El fichero mapeo_columnas.json no existe.")


    # borrar ficher csv
    if os.path.exists(file_dataBase_csv):
        os.remove(file_dataBase_csv)
        print(f"El fichero {file_dataBase_csv} ha sido borrado.")

    else:
        print(f"El fichero {file_dataBase_csv} no existe.")

    if os.path.exists(file_colores_csv):
        os.remove(file_colores_csv)
        print(f"El fichero {file_colores_csv} ha sido borrado.")
    else:
        print(f"El fichero {file_colores_csv} no existe.")

    # messagebox.showinfo("Procesamiento", "¡Procesamiento completado!")

    # al cerrar el messagebox, se cierra la ventana
    # root = tk.Tk()
    # root.withdraw()  # Ocultar la ventana principal
    # root.destroy()  # Cerrar la ventana principal
    # root.quit()  # Finalizar el programa




def limpiar(valor):
    if valor is None:
        return ""
    return str(valor).strip().replace("\u00A0", "").strip()

def analizar_excel(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.worksheets[0]

    # Leer fila 4
    fila4 = [cell.value for cell in ws[4] if isinstance(cell.value, (int, float))]

    if len(fila4) < 2:
        raise ValueError("La fila 4 no tiene suficientes números.")

    # Obtener mayor y segundo mayor
    nums = sorted(fila4, reverse=True)
    mayor, segundo = nums[0], nums[1]

    # Localizar columnas
    col_mayor = next(cell.column for cell in ws[4] if cell.value == mayor)
    col_segundo = next(cell.column for cell in ws[4] if cell.value == segundo)

    # Valores esperados SOLO en fila 1
    esperado_mayor = "MAIN INGREDIENT"
    esperado_segundo = "Property 3: PreparationMethod"

    def validar(col, esperado):
        valor_excel = limpiar(ws[1][col - 1].value)  # fila 1 → índice 1
        valor_esperado = limpiar(esperado)
        print(f"DEBUG: Col {col}: '{valor_excel}' vs '{valor_esperado}'")
        return valor_excel == valor_esperado

    # devuelve true si ambos son correctos, false si alguno de los dos no es correcto
    return validar(col_mayor, esperado_mayor) and validar(col_segundo, esperado_segundo)


# comprobamos que el kpi con mayor peso es el main ingredient de las columnas azules y el siguiente es el property 3: preparation method, y que ambos están en la fila 1 del excel de colores
if (analizar_excel(archivo2_nombre)):
    print("Los KPIs están correctamente identificados.")
    program()